#!/usr/bin/env python3
"""e-Stat 우나기(장어) 수입 통계 파일 구조 해부 (임시)."""
import re
import sys

import track_price as tp

BASE = "https://www.e-stat.go.jp"

for kind in ("0", "1"):
    url = f"{BASE}/stat-search/file-download?statInfId=000032208083&fileKind={kind}"
    print(f"\n########## DOWNLOAD fileKind={kind}")
    try:
        raw, final, ctype = tp.fetch(url)
    except Exception as e:
        print("  FAIL", e)
        continue
    print(f"  final={final}")
    print(f"  content-type={ctype} bytes={len(raw)} magic={raw[:8]!r}")
    path = f"/tmp/unagi_{kind}.bin"
    open(path, "wb").write(raw)

    if raw[:2] == b"PK":  # xlsx
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            print("  sheets:", wb.sheetnames)
            for name in wb.sheetnames[:3]:
                ws = wb[name]
                print(f"  --- sheet {name} dims={ws.max_row}x{ws.max_column}")
                for i, row in enumerate(ws.iter_rows(max_row=25, max_col=10, values_only=True)):
                    vals = [str(v)[:14] if v is not None else "" for v in row]
                    if any(vals):
                        print(f"   r{i+1}: {vals}")
                hits = 0
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    joined = " ".join(str(v) for v in row if v is not None)
                    if any(k in joined for k in ("中国", "中華", "China")):
                        print(f"   [中国 r{i+1}]: {[str(v)[:14] for v in row[:12]]}")
                        hits += 1
                        if hits >= 6:
                            break
        except Exception as e:
            print("  xlsx 파싱 실패:", e)
    elif raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":  # old xls
        try:
            import xlrd
            wb = xlrd.open_workbook(path)
            print("  sheets:", wb.sheet_names())
            ws = wb.sheet_by_index(0)
            print(f"  dims={ws.nrows}x{ws.ncols}")
            for i in range(min(25, ws.nrows)):
                vals = [str(ws.cell_value(i, c))[:14] for c in range(min(10, ws.ncols))]
                if any(vals):
                    print(f"   r{i+1}: {vals}")
            hits = 0
            for i in range(ws.nrows):
                joined = " ".join(str(ws.cell_value(i, c)) for c in range(ws.ncols))
                if any(k in joined for k in ("中国", "中華", "China")):
                    print(f"   [中国 r{i+1}]: {[str(ws.cell_value(i, c))[:14] for c in range(min(12, ws.ncols))]}")
                    hits += 1
                    if hits >= 6:
                        break
        except Exception as e:
            print("  xls 파싱 실패:", e)
    else:
        text = tp.decode_html(raw, ctype)
        print("  텍스트 미리보기:", re.sub(r"\s+", " ", text[:600]))
sys.exit(0)
